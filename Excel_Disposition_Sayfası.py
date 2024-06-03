import pandas as pd
from docx import Document
from openpyxl import load_workbook

# Word belgesinin dosya yolunu belirtin
doc_path = "C:/Users/alika/Desktop/Yeni Microsoft Word Belgesi.docx"
document = Document(doc_path)

# Excel belgesinin dosya yolunu belirtin
excel_path = "C:/Users/alika/Desktop/merged_data.xlsx"

# Belgedeki tüm tabloları alın
tables = document.tables

# Excel dosyasını oku ve tüm hücrelerdeki virgülleri noktaya çevir
df = pd.read_excel(excel_path, sheet_name='Sheet2').applymap(lambda x: str(x).replace(',', '.') if isinstance(x, str) else x)

# Sayısal değerlere dönüşüm yap
df = df.apply(pd.to_numeric, errors='ignore')

# İşlenecek verileri saklamak için bir liste oluşturun
results = []

# Tüm tabloları dolaşın
for table in tables:
    # Her bir tablodaki satırları dolaşın
    for row in table.rows:
        row_data = [cell.text for cell in row.cells]
        if row_data[0].startswith("KN"):
            kn_number = row_data[0].split("_")[0]
            formatted_output = "["+row_data[0]+"] " + row_data[1] + " (" + row_data[6] + ") "
            
            # Excel dosyasında eşleşen KN numarasını bulun
            matched_rows = df[df['Element'].str.startswith(kn_number)]

            for _, excel_row in matched_rows.iterrows():
                min_check = excel_row[4]
                max_check = excel_row[5]
                calculation1 = round((abs(excel_row[1]) - abs(excel_row[2])) - abs(excel_row[4]), 3)
                calculation2 = round(excel_row[5] - (abs(excel_row[1]) + abs(excel_row[3])), 3)

                if (abs(excel_row[1]) - abs(excel_row[2]) > min_check) and ((abs(excel_row[1]) + abs(excel_row[3])) < excel_row[5]):
                    updated_output = (
                        f"{formatted_output} checks min {min_check} max {max_check} or {calculation1} U/M or {calculation2} O/M."
                    )
                elif (abs(excel_row[1]) - abs(excel_row[2]) > min_check) and ((abs(excel_row[1]) + abs(excel_row[3])) > excel_row[5]):
                    updated_output = (
                        f"{formatted_output} checks min {min_check} or {calculation1} U/M."
                    )
                elif (abs(excel_row[2]) - abs(excel_row[3]) < min_check) and ((abs(excel_row[2]) + abs(excel_row[4])) < excel_row[5]):
                    updated_output = (
                        f"{formatted_output} checks max {max_check} or {calculation2} O/M."
                    )
                else:
                    updated_output = None  # Tolerans içerisinde, bir şey döndürmeye gerek yok
                
                if updated_output:
                    results.append(updated_output)

# Yeni bir DataFrame oluşturun
new_df = pd.DataFrame(results, columns=["Dimensional Flaw"])

# Gerekli diğer kolonları ekleyin
new_df["Disposition Type"] = ""
new_df["Serial Number"] = ""
new_df["Operation Number"] = ""
new_df["Cause OP. Number"] = ""
new_df["Quantity"] = ""
new_df["Cause Code"] = ""

# Excel dosyasını load_workbook ile yükleyin
wb = load_workbook(excel_path)

# Eğer DISPOSITION sayfası varsa silin
if "DISPOSITION" in wb.sheetnames:
    del wb["DISPOSITION"]

# Excel dosyasını kaydedin ve kapatın
wb.save(excel_path)
wb.close()

# Yeni bir sayfa ekleyin ve işlenen verileri yazın
with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a') as writer:
    new_df.to_excel(writer, sheet_name="DISPOSITION", index=False)

# Boş satırları DISPOSITION sayfasından silmek
wb = load_workbook(excel_path)
sheet = wb["DISPOSITION"]

# Boş satırları kontrol et ve sil
for row in sheet.iter_rows():
    if all(cell.value is None or cell.value == "" for cell in row):
        sheet.delete_rows(row[0].row, 1)

# Excel dosyasını tekrar kaydet
wb.save(excel_path)
wb.close()

print("Veriler başarıyla işlendi ve DISPOSITION sayfası tekrar oluşturuldu.")
